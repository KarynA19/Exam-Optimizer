from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from pydantic import ValidationError

from app.models.schedule import CourseImportResponse, CourseInput, DateRange, FixedExam, ScheduleProject, ValidationIssue
from app.services.validation import validate_project

COURSE_SHEET_NAME = "Courses"
FIXED_EXAMS_SHEET_NAME = "Fixed Exams"
EXAMPLE_PREFIX = "example"

COURSE_EXPECTED_HEADERS = [
    "course id",
    "course name",
    "semester",
    "is high failure",
    "prerequisites",
    "department",
]

COURSE_DISPLAY_HEADERS = [
    "Course ID",
    "Course Name",
    "Semester",
    "Is High Failure",
    "Prerequisites",
    "Department",
]

FIXED_EXAMS_EXPECTED_HEADERS = [
    "course id",
    "course name",
    "exam date",
    "prerequisites",
    "department",
]

FIXED_EXAMS_DISPLAY_HEADERS = [
    "Course ID",
    "Course Name",
    "Exam Date",
    "Prerequisites",
    "Department",
]

TRUTHY_VALUES = {"1", "true", "yes", "y"}
FALSY_VALUES = {"0", "false", "no", "n", ""}


class CourseImportError(Exception):
    def __init__(self, issues: list[ValidationIssue]):
        super().__init__(issues[0].message if issues else "Course import failed.")
        self.issues = issues


def _normalize_header(value: object) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _cell_text(value: object) -> str:
    return str(value or "").strip()


def _build_issue(message: str, row_number: int | None = None, course_code: str | None = None) -> ValidationIssue:
    row_prefix = f"Row {row_number}: " if row_number is not None else ""
    return ValidationIssue(
        code="invalid_import_file",
        severity="error",
        message=f"{row_prefix}{message}",
        related_course_code=course_code or None,
    )


def _build_header_issue(
    actual_headers: list[str],
    actual_header_labels: list[str],
    expected_headers: list[str],
    display_headers: list[str],
    sheet_name: str,
) -> ValidationIssue:
    mismatches: list[str] = []
    for index, (expected, actual) in enumerate(zip(expected_headers, actual_headers, strict=False), start=1):
        if expected == actual:
            continue
        column_name = chr(64 + index)
        actual_label = actual_header_labels[index - 1] or "blank"
        mismatches.append(f"Column {column_name} should be '{display_headers[index - 1]}' but is '{actual_label}'")

    if not mismatches:
        mismatches.append("the first row does not match the expected template")

    expected_order = ", ".join(f"{chr(65 + index)}: {header}" for index, header in enumerate(display_headers))
    return _build_issue(f"Template headers must match exactly in sheet '{sheet_name}'. Expected {expected_order}. Mismatch: {'; '.join(mismatches)}.")


def _build_duplicate_course_issues(row_numbers_by_code: dict[str, list[int]], sheet_name: str) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    for course_code, row_numbers in row_numbers_by_code.items():
        if len(row_numbers) < 2:
            continue
        joined_rows = ", ".join(str(row_number) for row_number in row_numbers)
        for row_number in row_numbers:
            issues.append(
                ValidationIssue(
                    code="duplicate_course_code",
                    severity="error",
                    message=f"Row {row_number}: Course ID '{course_code}' appears more than once in sheet '{sheet_name}' (rows {joined_rows}).",
                    related_course_code=course_code,
                )
            )
    return issues


def _build_missing_prerequisite_issues(
    rows: list[CourseInput | FixedExam],
    row_numbers_by_code: dict[str, list[int]],
    available_codes: set[str],
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    for course in rows:
        missing_codes = [
            prerequisite_code for prerequisite_code in course.prerequisite_course_codes if prerequisite_code not in available_codes
        ]
        if not missing_codes:
            continue

        row_number = row_numbers_by_code.get(course.course_code, [None])[0]
        issues.append(
            ValidationIssue(
                code="missing_prerequisite_target",
                severity="error",
                message=(
                    f"Row {row_number}: Prerequisites references {', '.join(repr(code) for code in missing_codes)}, "
                    "but those Course ID values do not exist in this file."
                )
                if row_number is not None
                else (
                    f"Prerequisites references {', '.join(repr(code) for code in missing_codes)}, "
                    "but those Course ID values do not exist in this file."
                ),
                related_course_code=course.course_code,
            )
        )
    return issues


def _contextualize_validation_issues(
    issues: list[ValidationIssue],
    row_numbers_by_code: dict[str, list[int]],
) -> list[ValidationIssue]:
    contextualized: list[ValidationIssue] = []
    for issue in issues:
        row_number = row_numbers_by_code.get(issue.related_course_code or "", [None])[0]
        if row_number is None:
            contextualized.append(issue)
            continue
        contextualized.append(
            issue.model_copy(
                update={
                    "message": f"Row {row_number}: {issue.message}",
                }
            )
        )
    return contextualized


def _parse_semester(value: object) -> int:
    if isinstance(value, bool):
        raise ValueError("Semester must be a number between 1 and 12.")
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)

    text = _cell_text(value)
    if not text:
        raise ValueError("Semester is required.")

    try:
        return int(text)
    except ValueError as error:
        raise ValueError("Semester must be a number between 1 and 12.") from error


def _parse_high_failure(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and value in {0, 1}:
        return bool(value)
    if isinstance(value, float) and value in {0.0, 1.0}:
        return bool(int(value))

    normalized = _cell_text(value).lower()
    if normalized in TRUTHY_VALUES:
        return True
    if normalized in FALSY_VALUES:
        return False

    raise ValueError("Is High Failure must be one of yes/no, true/false, or 1/0.")


def _parse_department(value: object) -> str | None:
    text = _cell_text(value).upper()
    if not text:
        return None
    if text in {"SW", "IS"}:
        return text
    raise ValueError("Department must be empty, SW, or IS.")


def _parse_exam_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _cell_text(value)
    if not text:
        raise ValueError("Exam Date is required.")

    for fmt in ("%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    raise ValueError("Exam Date must be in YYYY-MM-DD format.")


def _is_example_row(raw_values: list[object]) -> bool:
    first_cell = _cell_text(raw_values[0]).lower()
    return first_cell.startswith(EXAMPLE_PREFIX)


def export_course_template_excel() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = COURSE_SHEET_NAME
    worksheet.append(COURSE_DISPLAY_HEADERS)
    worksheet.append(["EXAMPLE_CS101", "Example Course", 1, "no", "", "SW"])

    fixed_exams_worksheet = workbook.create_sheet(FIXED_EXAMS_SHEET_NAME)
    fixed_exams_worksheet.append(FIXED_EXAMS_DISPLAY_HEADERS)
    fixed_exams_worksheet.append(["EXAMPLE_CS101", "Example Course", "2026-06-20", "", "SW"])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def import_courses_from_excel(content: bytes) -> CourseImportResponse:
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as error:  # pragma: no cover - exact openpyxl error type is not stable
        raise CourseImportError([_build_issue("Uploaded file is not a valid .xlsx workbook.")]) from error

    if COURSE_SHEET_NAME not in workbook.sheetnames:
        raise CourseImportError([_build_issue(f"Workbook must include a '{COURSE_SHEET_NAME}' sheet.")])
    if FIXED_EXAMS_SHEET_NAME not in workbook.sheetnames:
        raise CourseImportError([_build_issue(f"Workbook must include a '{FIXED_EXAMS_SHEET_NAME}' sheet.")])

    worksheet = workbook[COURSE_SHEET_NAME]
    fixed_exams_worksheet = workbook[FIXED_EXAMS_SHEET_NAME]

    course_header_labels = [_cell_text(worksheet.cell(row=1, column=column).value) for column in range(1, 7)]
    course_headers = [_normalize_header(header) for header in course_header_labels]
    if course_headers != COURSE_EXPECTED_HEADERS:
        raise CourseImportError([
            _build_header_issue(
                course_headers,
                course_header_labels,
                COURSE_EXPECTED_HEADERS,
                COURSE_DISPLAY_HEADERS,
                COURSE_SHEET_NAME,
            )
        ])

    fixed_header_labels = [_cell_text(fixed_exams_worksheet.cell(row=1, column=column).value) for column in range(1, 6)]
    fixed_headers = [_normalize_header(header) for header in fixed_header_labels]
    if fixed_headers != FIXED_EXAMS_EXPECTED_HEADERS:
        raise CourseImportError([
            _build_header_issue(
                fixed_headers,
                fixed_header_labels,
                FIXED_EXAMS_EXPECTED_HEADERS,
                FIXED_EXAMS_DISPLAY_HEADERS,
                FIXED_EXAMS_SHEET_NAME,
            )
        ])

    issues: list[ValidationIssue] = []
    courses: list[CourseInput] = []
    fixed_exams: list[FixedExam] = []
    row_numbers_by_code: dict[str, list[int]] = {}
    row_numbers_by_fixed_exam_code: dict[str, list[int]] = {}

    for row_number in range(2, worksheet.max_row + 1):
        raw_values = [worksheet.cell(row=row_number, column=column).value for column in range(1, 7)]
        if all(_cell_text(value) == "" for value in raw_values):
            continue
        if _is_example_row(raw_values):
            continue

        course_code = _cell_text(raw_values[0])
        course_name = _cell_text(raw_values[1])
        prerequisite_codes = _cell_text(raw_values[4])

        try:
            semester_number = _parse_semester(raw_values[2])
            high_failure_rate = _parse_high_failure(raw_values[3])
            department = _parse_department(raw_values[5])
            courses.append(
                CourseInput(
                    course_code=course_code,
                    course_name=course_name,
                    semester_number=semester_number,
                    high_failure_rate=high_failure_rate,
                    department=department,
                    prerequisite_course_codes=prerequisite_codes,
                )
            )
            row_numbers_by_code.setdefault(course_code, []).append(row_number)
        except ValueError as error:
            issues.append(_build_issue(str(error), row_number=row_number, course_code=course_code or None))
        except ValidationError as error:
            for validation_error in error.errors():
                issues.append(_build_issue(validation_error["msg"], row_number=row_number, course_code=course_code or None))

    for row_number in range(2, fixed_exams_worksheet.max_row + 1):
        raw_values = [fixed_exams_worksheet.cell(row=row_number, column=column).value for column in range(1, 6)]
        if all(_cell_text(value) == "" for value in raw_values):
            continue
        if _is_example_row(raw_values):
            continue

        course_code = _cell_text(raw_values[0])
        course_name = _cell_text(raw_values[1])
        prerequisite_codes = _cell_text(raw_values[3])

        try:
            exam_date = _parse_exam_date(raw_values[2])
            department = _parse_department(raw_values[4])
            fixed_exams.append(
                FixedExam(
                    course_code=course_code,
                    course_name=course_name,
                    prerequisite_course_codes=prerequisite_codes,
                    exam_date=exam_date,
                    locked=True,
                    department=department,
                )
            )
            row_numbers_by_fixed_exam_code.setdefault(course_code, []).append(row_number)
        except ValueError as error:
            issues.append(_build_issue(str(error), row_number=row_number, course_code=course_code or None))
        except ValidationError as error:
            for validation_error in error.errors():
                issues.append(_build_issue(validation_error["msg"], row_number=row_number, course_code=course_code or None))

    if not courses and not fixed_exams and not issues:
        issues.append(_build_issue("The workbook does not contain any importable rows in Courses or Fixed Exams sheets."))

    issues.extend(_build_duplicate_course_issues(row_numbers_by_code, COURSE_SHEET_NAME))
    issues.extend(_build_duplicate_course_issues(row_numbers_by_fixed_exam_code, FIXED_EXAMS_SHEET_NAME))

    available_codes = set(row_numbers_by_code) | set(row_numbers_by_fixed_exam_code)
    issues.extend(_build_missing_prerequisite_issues(courses, row_numbers_by_code, available_codes))
    issues.extend(_build_missing_prerequisite_issues(fixed_exams, row_numbers_by_fixed_exam_code, available_codes))

    if issues:
        raise CourseImportError(issues)

    validation_project = ScheduleProject(
        project_name="Imported Courses",
        moed_windows=[DateRange(start_date=date(2026, 1, 1), end_date=date(2026, 12, 31))],
        courses=courses,
        fixed_exams=fixed_exams,
    )
    validation_issues = validate_project(validation_project)

    if validation_issues:
        contextual_row_numbers = {**row_numbers_by_fixed_exam_code, **row_numbers_by_code}
        raise CourseImportError(_contextualize_validation_issues(validation_issues, contextual_row_numbers))

    return CourseImportResponse(
        imported_count=len(courses),
        courses=courses,
        fixed_exams_imported_count=len(fixed_exams),
        fixed_exams=fixed_exams,
    )