from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.services.course_import import CourseImportError, export_course_template_excel, import_courses_from_excel


def build_workbook_bytes(
    rows: list[list[object]],
    headers: list[str] | None = None,
    fixed_rows: list[list[object]] | None = None,
    fixed_headers: list[str] | None = None,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Courses"
    worksheet.append(headers or ["Course ID", "Course Name", "Semester", "Is High Failure", "Prerequisites", "Department"])
    for row in rows:
        worksheet.append(row)

    fixed_worksheet = workbook.create_sheet("Fixed Exams")
    fixed_worksheet.append(fixed_headers or ["Course ID", "Course Name", "Semester", "Exam Date", "Prerequisites", "Department"])
    for row in fixed_rows or []:
        fixed_worksheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_import_courses_from_excel_parses_valid_template() -> None:
    content = build_workbook_bytes(
        [
            ["ALG1", "Algebra 1", 1, "yes", "", ""],
            ["ALG2", "Algebra 2", 2, "no", "ALG1", "sw"],
        ]
    )

    result = import_courses_from_excel(content)

    assert result.imported_count == 2
    assert result.courses[0].course_code == "ALG1"
    assert result.courses[0].high_failure_rate is True
    assert result.courses[0].department is None
    assert result.courses[1].prerequisite_course_codes == ["ALG1"]
    assert result.courses[1].department == "SW"
    assert result.fixed_exams_imported_count == 0
    assert result.fixed_exams == []


def test_import_courses_from_excel_parses_multiple_prerequisites() -> None:
    content = build_workbook_bytes(
        [
            ["ALG1", "Algebra 1", 1, "yes", "", ""],
            ["CALC1", "Calculus 1", 2, "no", "", "IS"],
            ["ALG3", "Advanced Algebra", 3, "no", "ALG1, CALC1", ""],
        ]
    )

    result = import_courses_from_excel(content)

    assert result.courses[2].prerequisite_course_codes == ["ALG1", "CALC1"]


def test_import_courses_from_excel_rejects_invalid_headers() -> None:
    content = build_workbook_bytes(
        [["ALG1", "Algebra 1", 1, "yes", "", ""]],
        headers=["Wrong", "Course Name", "Semester", "Is High Failure", "Prerequisites", "Department"],
    )

    try:
        import_courses_from_excel(content)
    except CourseImportError as error:
        assert any("Column A should be 'Course ID'" in issue.message for issue in error.issues)
    else:
        raise AssertionError("Expected CourseImportError for invalid headers.")


def test_import_courses_from_excel_rejects_invalid_fixed_exam_headers() -> None:
    content = build_workbook_bytes(
        [["ALG1", "Algebra 1", 1, "yes", "", ""]],
        fixed_headers=["Wrong", "Course Name", "Semester", "Exam Date", "Prerequisites", "Department"],
    )

    try:
        import_courses_from_excel(content)
    except CourseImportError as error:
        assert any("sheet 'Fixed Exams'" in issue.message for issue in error.issues)
    else:
        raise AssertionError("Expected CourseImportError for invalid fixed exam headers.")


def test_import_courses_from_excel_rejects_invalid_high_failure_value() -> None:
    content = build_workbook_bytes([["ALG1", "Algebra 1", 1, "maybe", "", ""]])

    try:
        import_courses_from_excel(content)
    except CourseImportError as error:
        assert any("Row 2: Is High Failure must be one of yes/no, true/false, or 1/0." == issue.message for issue in error.issues)
    else:
        raise AssertionError("Expected CourseImportError for invalid high-failure value.")


def test_import_courses_from_excel_rejects_missing_prerequisite_target() -> None:
    content = build_workbook_bytes([["ALG2", "Algebra 2", 2, "no", "ALG1", ""]])

    try:
        import_courses_from_excel(content)
    except CourseImportError as error:
        assert any("Row 2: Prerequisites references 'ALG1'" in issue.message for issue in error.issues)
    else:
        raise AssertionError("Expected CourseImportError for missing prerequisite target.")


def test_import_courses_from_excel_rejects_duplicate_course_codes() -> None:
    content = build_workbook_bytes(
        [
            ["ALG1", "Algebra 1", 1, "yes", "", ""],
            ["ALG1", "Linear Algebra", 1, "no", "", "IS"],
        ]
    )

    try:
        import_courses_from_excel(content)
    except CourseImportError as error:
        assert any("Row 2: Course ID 'ALG1' appears more than once" in issue.message for issue in error.issues)
    else:
        raise AssertionError("Expected CourseImportError for duplicate course codes.")


def test_import_courses_from_excel_rejects_empty_workbook() -> None:
    content = build_workbook_bytes([])

    try:
        import_courses_from_excel(content)
    except CourseImportError as error:
        assert any("does not contain any importable rows" in issue.message for issue in error.issues)
    else:
        raise AssertionError("Expected CourseImportError for empty workbook.")


def test_import_courses_from_excel_rejects_invalid_department_value() -> None:
    content = build_workbook_bytes([["ALG1", "Algebra 1", 1, "yes", "", "EE"]])

    try:
        import_courses_from_excel(content)
    except CourseImportError as error:
        assert any("Row 2: Department must be empty, SW, or IS." == issue.message for issue in error.issues)
    else:
        raise AssertionError("Expected CourseImportError for invalid department value.")


def test_export_course_template_excel_uses_current_headers() -> None:
    content = export_course_template_excel()

    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    courses_sheet = workbook["Courses"]
    fixed_sheet = workbook["Fixed Exams"]

    course_headers = [courses_sheet.cell(row=1, column=column).value for column in range(1, 7)]
    fixed_headers = [fixed_sheet.cell(row=1, column=column).value for column in range(1, 7)]

    assert course_headers == ["Course ID", "Course Name", "Semester", "Is High Failure", "Prerequisites", "Department"]
    assert fixed_headers == ["Course ID", "Course Name", "Semester", "Exam Date", "Prerequisites", "Department"]
    assert str(courses_sheet.cell(row=2, column=1).value).startswith("EXAMPLE")
    assert str(fixed_sheet.cell(row=2, column=1).value).startswith("EXAMPLE")


def test_import_courses_from_excel_parses_fixed_exams_and_skips_example_rows() -> None:
    content = build_workbook_bytes(
        [
            ["EXAMPLE_CS101", "Example Course", 1, "no", "", "SW"],
            ["ALG1", "Algebra 1", 1, "yes", "", ""],
        ],
        fixed_rows=[
            ["EXAMPLE_CS101", "Example Course", 1, "2026-06-22", "", "SW"],
            ["ALG1", "Algebra 1", 1, "2026-06-22", "", "SW"],
        ],
    )

    result = import_courses_from_excel(content)

    assert result.imported_count == 1
    assert result.fixed_exams_imported_count == 1
    assert result.courses[0].course_code == "ALG1"
    assert result.fixed_exams[0].course_code == "ALG1"
    assert result.fixed_exams[0].semester_number == 1